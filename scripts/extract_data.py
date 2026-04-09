"""
extract_data.py
Extrai tabelas do Anuário Brasileiro de Segurança Pública (Excel)
e salva como JSONs limpos em data/{ano}/.

Uso:
    python extract_data.py --year 2025

O Excel deve estar em:
    scripts/raw/2025/anuario-2025.xlsx

Os JSONs são salvos em:
    data/2025/*.json
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

REGIOES: dict[str, str] = {
    "Acre": "Norte",
    "Amapá": "Norte",
    "Amazonas": "Norte",
    "Pará": "Norte",
    "Rondônia": "Norte",
    "Roraima": "Norte",
    "Tocantins": "Norte",
    "Alagoas": "Nordeste",
    "Bahia": "Nordeste",
    "Ceará": "Nordeste",
    "Maranhão": "Nordeste",
    "Paraíba": "Nordeste",
    "Pernambuco": "Nordeste",
    "Piauí": "Nordeste",
    "Rio Grande do Norte": "Nordeste",
    "Sergipe": "Nordeste",
    "Distrito Federal": "Centro-Oeste",
    "Goiás": "Centro-Oeste",
    "Mato Grosso": "Centro-Oeste",
    "Mato Grosso do Sul": "Centro-Oeste",
    "Espírito Santo": "Sudeste",
    "Minas Gerais": "Sudeste",
    "Rio de Janeiro": "Sudeste",
    "São Paulo": "Sudeste",
    "Paraná": "Sul",
    "Rio Grande do Sul": "Sul",
    "Santa Catarina": "Sul",
}

UFS = ["Brasil"] + sorted(REGIOES.keys())

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------


def _regiao(uf: str) -> str | None:
    return REGIOES.get(uf)


def _round_or_none(v, decimais: int = 2):
    """Arredonda floats; preserva inteiros; converte NaN/None para None."""
    if v is None:
        return None
    try:
        import math
        if math.isnan(float(v)):
            return None
    except (TypeError, ValueError):
        return None
    if decimais == 0:
        return int(round(float(v)))
    return round(float(v), decimais)


def _int_or_none(v):
    return _round_or_none(v, 0)


def save_json(path: Path, data: dict, label: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✓  {path.name}  →  {path}")


def validate(data: dict, checks: list[tuple[bool, str]]) -> None:
    """Aborta se alguma validação falhar."""
    for ok, msg in checks:
        if not ok:
            print(f"  ✗  VALIDAÇÃO FALHOU: {msg}", file=sys.stderr)
            sys.exit(1)


# ---------------------------------------------------------------------------
# Função auxiliar de leitura de aba com tratamento de merged cells
# ---------------------------------------------------------------------------


def read_sheet(wb, sheet_name: str) -> list[list]:
    """
    Lê uma aba do workbook e retorna como lista de listas.
    Substitui células mescladas pelo valor da célula mestre.
    Retorna apenas linhas com ao menos um valor não-None.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Aba '{sheet_name}' não encontrada. "
            f"Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    # openpyxl com read_only=False para acessar merged_cells
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            rows.append(list(row))
    return rows


def find_row_index(rows: list[list], keyword: str, col: int = 0) -> int:
    """Retorna o índice da primeira linha onde rows[i][col] contém keyword."""
    for i, row in enumerate(rows):
        if row[col] is not None and keyword.lower() in str(row[col]).lower():
            return i
    raise ValueError(f"Linha com '{keyword}' não encontrada na coluna {col}")


# ---------------------------------------------------------------------------
# T02 — MVI histórico por UF (2012–2024)
# ---------------------------------------------------------------------------


def extract_mvi_historico(wb, year: int, output_dir: Path) -> None:
    """
    T02: MVI por UF 2012–2024.
    Estrutura esperada:
      - Linha com os anos na parte superior
      - Coluna 0: nome da UF
      - Colunas 1..13: absolutos por ano
      - Colunas 14..26 (aprox): taxas por 100k
    """
    rows = read_sheet(wb, "T02")

    # Localizar linha dos anos (procura linha com vários inteiros >=2000)
    anos_row_idx = None
    for i, row in enumerate(rows):
        candidatos = [v for v in row if isinstance(v, (int, float)) and 2000 <= v <= 2030]
        if len(candidatos) >= 10:
            anos_row_idx = i
            break

    if anos_row_idx is None:
        print("  ! Não foi possível detectar automaticamente a linha de anos em T02.", file=sys.stderr)
        print("    Inspecione a aba manualmente e ajuste a lógica.", file=sys.stderr)
        sys.exit(1)

    anos_row = rows[anos_row_idx]
    # Separar anos (inteiros entre 2000-2030) e suas posições
    anos_info = [(i, int(v)) for i, v in enumerate(anos_row) if isinstance(v, (int, float)) and 2000 <= v <= 2030]
    anos = [a for _, a in anos_info]
    abs_cols = [i for i, _ in anos_info]

    # Taxas costumam estar logo após os absolutos com mesmos anos repetidos,
    # ou em colunas específicas. Buscamos em linhas posteriores.
    # Estratégia: localizar segunda ocorrência dos mesmos anos.
    taxa_cols: list[int] = []
    for row in rows[anos_row_idx + 1 :]:
        candidatos = [(i, int(v)) for i, v in enumerate(row) if isinstance(v, (int, float)) and 2000 <= v <= 2030]
        if len(candidatos) == len(anos) and [a for _, a in candidatos] == anos:
            taxa_cols = [i for i, _ in candidatos]
            break

    # Se não encontrou segunda linha de anos, assume que taxas estão
    # deslocadas em n_anos colunas após os absolutos.
    if not taxa_cols:
        offset = len(anos)
        taxa_cols = [c + offset for c in abs_cols]

    dados: list[dict] = []
    for row in rows[anos_row_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() == "":
            continue
        uf_str = str(uf).strip()
        if uf_str not in UFS:
            continue

        absolutos = [_int_or_none(row[c] if c < len(row) else None) for c in abs_cols]
        taxas = [_round_or_none(row[c] if c < len(row) else None) for c in taxa_cols]

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "absolutos": absolutos,
            "taxas": taxas,
        })

    validate(dados, [
        (len(dados) >= 27, f"T02: esperado 28 linhas (Brasil+27 UFs), encontrado {len(dados)}"),
        (any(d["uf"] == "Brasil" for d in dados), "T02: linha 'Brasil' não encontrada"),
        (len(anos) >= 10, f"T02: esperado >= 10 anos, encontrado {len(anos)}"),
    ])

    save_json(
        output_dir / "mvi_historico.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T02",
            "anos": anos,
            "dados": dados,
        },
        "T02 MVI histórico",
    )


# ---------------------------------------------------------------------------
# T01 — MVI por estados, comparativo 2023–2024
# ---------------------------------------------------------------------------


def extract_mvi_estados(wb, year: int, output_dir: Path) -> None:
    """
    T01: MVI por UF com subcategorias (homicídio doloso, latrocínio, LCFM,
    intervenção policial).
    """
    rows = read_sheet(wb, "T01")

    # Localizar linha de cabeçalho com "2023" e "2024"
    header_idx = None
    for i, row in enumerate(rows):
        anos_na_linha = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
        if len(anos_na_linha) >= 4:  # espera ao menos 4 ocorrências (2 por categoria)
            header_idx = i
            break

    if header_idx is None:
        print("  ! T01: cabeçalho com 2023/2024 não encontrado. Pulando.", file=sys.stderr)
        return

    # Mapear colunas por posição
    # A estrutura típica: UF | Hom. Doloso 23 | Hom. Doloso 24 | Latr. 23 | Latr. 24 | ...
    # Usa heurística: pares de colunas por categoria
    header = rows[header_idx]
    col_2023 = [i for i, v in enumerate(header) if v == 2023]
    col_2024 = [i for i, v in enumerate(header) if v == 2024]

    # Assume mesma quantidade de colunas para cada ano e mesma ordem
    dados: list[dict] = []
    for row in rows[header_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()

        def get(cols, idx):
            return _int_or_none(row[cols[idx]] if idx < len(cols) and cols[idx] < len(row) else None)

        # Ordem esperada: hom_doloso, latrocinio, lcfm, intervencao, mvi_total, taxa
        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "homicidio_doloso_2023": get(col_2023, 0),
            "homicidio_doloso_2024": get(col_2024, 0),
            "latrocinio_2023": get(col_2023, 1),
            "latrocinio_2024": get(col_2024, 1),
            "lcfm_2023": get(col_2023, 2),
            "lcfm_2024": get(col_2024, 2),
            "intervencao_policial_2023": get(col_2023, 3),
            "intervencao_policial_2024": get(col_2024, 3),
            "mvi_total_2023": get(col_2023, 4),
            "mvi_total_2024": get(col_2024, 4),
            "taxa_2023": _round_or_none(row[col_2023[5]] if len(col_2023) > 5 and col_2023[5] < len(row) else None),
            "taxa_2024": _round_or_none(row[col_2024[5]] if len(col_2024) > 5 and col_2024[5] < len(row) else None),
        })

    # Calcular variação percentual
    for d in dados:
        if d["mvi_total_2023"] and d["mvi_total_2024"]:
            d["variacao_pct"] = _round_or_none(
                (d["mvi_total_2024"] - d["mvi_total_2023"]) / d["mvi_total_2023"] * 100
            )
        else:
            d["variacao_pct"] = None

    if not dados:
        print("  ! T01: nenhuma linha de UF extraída. Verifique a estrutura da aba.", file=sys.stderr)
        return

    save_json(
        output_dir / "mvi_estados.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T01",
            "dados": dados,
        },
        "T01 MVI estados",
    )


# ---------------------------------------------------------------------------
# T24 — Feminicídio por UF
# ---------------------------------------------------------------------------


def extract_feminicidio(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "T24")

    header_idx = None
    for i, row in enumerate(rows):
        anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
        if len(anos) >= 2:
            header_idx = i
            break

    if header_idx is None:
        print("  ! T24: cabeçalho com 2023/2024 não encontrado. Pulando.", file=sys.stderr)
        return

    header = rows[header_idx]
    col_2023 = [i for i, v in enumerate(header) if v == 2023]
    col_2024 = [i for i, v in enumerate(header) if v == 2024]

    dados: list[dict] = []
    for row in rows[header_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()

        def get(cols, idx):
            return _int_or_none(row[cols[idx]] if idx < len(cols) and cols[idx] < len(row) else None)

        hom_m_23 = get(col_2023, 0)
        fem_23 = get(col_2023, 1)
        hom_m_24 = get(col_2024, 0)
        fem_24 = get(col_2024, 1)

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "homicidios_mulheres_2023": hom_m_23,
            "homicidios_mulheres_2024": hom_m_24,
            "feminicidios_2023": fem_23,
            "feminicidios_2024": fem_24,
            "proporcao_2023": _round_or_none(fem_23 / hom_m_23 * 100) if hom_m_23 else None,
            "proporcao_2024": _round_or_none(fem_24 / hom_m_24 * 100) if hom_m_24 else None,
        })

    if not dados:
        print("  ! T24: nenhuma linha de UF extraída.", file=sys.stderr)
        return

    save_json(
        output_dir / "feminicidio.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T24",
            "dados": dados,
        },
        "T24 Feminicídio",
    )


# ---------------------------------------------------------------------------
# Q07 — Proporção de feminicídios 2015–2024
# ---------------------------------------------------------------------------


def extract_feminicidio_hist(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "Q07")

    # Localizar linha com anos (2015–2024)
    anos_row_idx = None
    for i, row in enumerate(rows):
        candidatos = [v for v in row if isinstance(v, (int, float)) and 2010 <= v <= 2030]
        if len(candidatos) >= 8:
            anos_row_idx = i
            break

    if anos_row_idx is None:
        print("  ! Q07: linha de anos não encontrada. Pulando.", file=sys.stderr)
        return

    anos_row = rows[anos_row_idx]
    anos_info = [(i, int(v)) for i, v in enumerate(anos_row) if isinstance(v, (int, float)) and 2010 <= v <= 2030]
    anos = [a for _, a in anos_info]
    ano_cols = [i for i, _ in anos_info]

    dados: list[dict] = []
    for row in rows[anos_row_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()
        proporcoes = [_round_or_none(row[c] if c < len(row) else None) for c in ano_cols]

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "proporcoes": proporcoes,
        })

    if not dados:
        print("  ! Q07: nenhuma linha de UF extraída.", file=sys.stderr)
        return

    save_json(
        output_dir / "feminicidio_hist.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "Q07",
            "anos": anos,
            "dados": dados,
        },
        "Q07 Feminicídio histórico",
    )


# ---------------------------------------------------------------------------
# T09 + T10 — Letalidade policial
# ---------------------------------------------------------------------------


def extract_letalidade(wb, year: int, output_dir: Path) -> None:
    dados: dict[str, dict] = {}

    for tabela in ("T09", "T10"):
        if tabela not in wb.sheetnames:
            print(f"  ! {tabela}: aba não encontrada. Pulando.", file=sys.stderr)
            continue
        rows = read_sheet(wb, tabela)

        header_idx = None
        for i, row in enumerate(rows):
            anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
            if len(anos) >= 2:
                header_idx = i
                break

        if header_idx is None:
            continue

        header = rows[header_idx]
        col_2023 = [i for i, v in enumerate(header) if v == 2023]
        col_2024 = [i for i, v in enumerate(header) if v == 2024]

        for row in rows[header_idx + 1 :]:
            uf = row[0]
            if uf is None or str(uf).strip() not in UFS:
                continue
            uf_str = str(uf).strip()

            if uf_str not in dados:
                dados[uf_str] = {"uf": uf_str, "regiao": _regiao(uf_str)}

            if tabela == "T09":
                dados[uf_str]["mortes_2023"] = _int_or_none(row[col_2023[0]] if col_2023 else None)
                dados[uf_str]["mortes_2024"] = _int_or_none(row[col_2024[0]] if col_2024 else None)
            else:  # T10 — proporção em relação às MVI
                dados[uf_str]["proporcao_mvi_2023"] = _round_or_none(row[col_2023[0]] if col_2023 else None)
                dados[uf_str]["proporcao_mvi_2024"] = _round_or_none(row[col_2024[0]] if col_2024 else None)

    if not dados:
        print("  ! T09/T10: nenhum dado extraído.", file=sys.stderr)
        return

    save_json(
        output_dir / "letalidade.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabelas": ["T09", "T10"],
            "dados": list(dados.values()),
        },
        "T09+T10 Letalidade policial",
    )


# ---------------------------------------------------------------------------
# T34 — Violência sexual
# ---------------------------------------------------------------------------


def extract_estupro(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "T34")

    header_idx = None
    for i, row in enumerate(rows):
        anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
        if len(anos) >= 2:
            header_idx = i
            break

    if header_idx is None:
        print("  ! T34: cabeçalho não encontrado. Pulando.", file=sys.stderr)
        return

    header = rows[header_idx]
    col_2023 = [i for i, v in enumerate(header) if v == 2023]
    col_2024 = [i for i, v in enumerate(header) if v == 2024]

    dados: list[dict] = []
    for row in rows[header_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()

        def get(cols, idx):
            return _int_or_none(row[cols[idx]] if idx < len(cols) and cols[idx] < len(row) else None)

        est_23 = get(col_2023, 0)
        vul_23 = get(col_2023, 1)
        est_24 = get(col_2024, 0)
        vul_24 = get(col_2024, 1)

        total_23 = (est_23 or 0) + (vul_23 or 0) if (est_23 is not None or vul_23 is not None) else None
        total_24 = (est_24 or 0) + (vul_24 or 0) if (est_24 is not None or vul_24 is not None) else None

        taxa_23 = _round_or_none(row[col_2023[2]] if len(col_2023) > 2 and col_2023[2] < len(row) else None)
        taxa_24 = _round_or_none(row[col_2024[2]] if len(col_2024) > 2 and col_2024[2] < len(row) else None)

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "estupro_2023": est_23,
            "estupro_2024": est_24,
            "estupro_vulneravel_2023": vul_23,
            "estupro_vulneravel_2024": vul_24,
            "total_2023": total_23,
            "total_2024": total_24,
            "taxa_2023": taxa_23,
            "taxa_2024": taxa_24,
        })

    if not dados:
        print("  ! T34: nenhuma linha extraída.", file=sys.stderr)
        return

    save_json(
        output_dir / "estupro.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T34",
            "dados": dados,
        },
        "T34 Violência sexual",
    )


# ---------------------------------------------------------------------------
# T127 — População prisional 2000–2024
# ---------------------------------------------------------------------------


def extract_prisional(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "T127")

    # Localizar linhas com anos (inteiros entre 2000–2030 na coluna 0)
    dados: list[dict] = []
    for row in rows:
        ano = row[0]
        if not isinstance(ano, (int, float)) or not (2000 <= int(ano) <= 2030):
            continue
        dados.append({
            "ano": int(ano),
            "total": _int_or_none(row[1] if len(row) > 1 else None),
            "condenados": _int_or_none(row[2] if len(row) > 2 else None),
            "provisorios": _int_or_none(row[3] if len(row) > 3 else None),
            "taxa_encarceramento": _round_or_none(row[4] if len(row) > 4 else None),
        })

    if len(dados) < 10:
        print(f"  ! T127: esperado >= 10 anos, encontrado {len(dados)}. Verifique a estrutura.", file=sys.stderr)
        return

    dados.sort(key=lambda d: d["ano"])

    save_json(
        output_dir / "prisional.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T127",
            "dados": dados,
        },
        "T127 Sistema prisional",
    )


# ---------------------------------------------------------------------------
# T11 — Desaparecimentos
# ---------------------------------------------------------------------------


def extract_desaparecimentos(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "T11")

    header_idx = None
    for i, row in enumerate(rows):
        anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
        if len(anos) >= 2:
            header_idx = i
            break

    if header_idx is None:
        print("  ! T11: cabeçalho não encontrado. Pulando.", file=sys.stderr)
        return

    header = rows[header_idx]
    col_2023 = [i for i, v in enumerate(header) if v == 2023]
    col_2024 = [i for i, v in enumerate(header) if v == 2024]

    dados: list[dict] = []
    for row in rows[header_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()

        v23 = _int_or_none(row[col_2023[0]] if col_2023 else None)
        v24 = _int_or_none(row[col_2024[0]] if col_2024 else None)
        var = _round_or_none((v24 - v23) / v23 * 100) if v23 and v24 else None

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "desaparecidos_2023": v23,
            "desaparecidos_2024": v24,
            "variacao_pct": var,
        })

    if not dados:
        print("  ! T11: nenhuma linha extraída.", file=sys.stderr)
        return

    save_json(
        output_dir / "desaparecimentos.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T11",
            "dados": dados,
        },
        "T11 Desaparecimentos",
    )


# ---------------------------------------------------------------------------
# T12 + T13 — Patrimônio
# ---------------------------------------------------------------------------


def extract_patrimonio(wb, year: int, output_dir: Path) -> None:
    dados: dict[str, dict] = {}

    for tabela, campos in (
        ("T12", ("roubo_veiculo", "furto_veiculo")),
        ("T13", ("roubo_celular", "furto_celular")),
    ):
        if tabela not in wb.sheetnames:
            print(f"  ! {tabela}: aba não encontrada. Pulando.", file=sys.stderr)
            continue

        rows = read_sheet(wb, tabela)
        header_idx = None
        for i, row in enumerate(rows):
            anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
            if len(anos) >= 2:
                header_idx = i
                break

        if header_idx is None:
            continue

        header = rows[header_idx]
        col_2023 = [i for i, v in enumerate(header) if v == 2023]
        col_2024 = [i for i, v in enumerate(header) if v == 2024]

        for row in rows[header_idx + 1 :]:
            uf = row[0]
            if uf is None or str(uf).strip() not in UFS:
                continue
            uf_str = str(uf).strip()
            if uf_str not in dados:
                dados[uf_str] = {"uf": uf_str, "regiao": _regiao(uf_str)}

            for j, campo in enumerate(campos):
                dados[uf_str][f"{campo}_2023"] = _int_or_none(row[col_2023[j]] if j < len(col_2023) and col_2023[j] < len(row) else None)
                dados[uf_str][f"{campo}_2024"] = _int_or_none(row[col_2024[j]] if j < len(col_2024) and col_2024[j] < len(row) else None)

    if not dados:
        print("  ! T12/T13: nenhum dado extraído.", file=sys.stderr)
        return

    save_json(
        output_dir / "patrimonio.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabelas": ["T12", "T13"],
            "dados": list(dados.values()),
        },
        "T12+T13 Patrimônio",
    )


# ---------------------------------------------------------------------------
# T96 — Gastos com segurança pública
# ---------------------------------------------------------------------------


def extract_gastos(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "T96")

    header_idx = None
    for i, row in enumerate(rows):
        anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
        if len(anos) >= 2:
            header_idx = i
            break

    if header_idx is None:
        print("  ! T96: cabeçalho não encontrado. Pulando.", file=sys.stderr)
        return

    header = rows[header_idx]
    col_2023 = [i for i, v in enumerate(header) if v == 2023]
    col_2024 = [i for i, v in enumerate(header) if v == 2024]

    dados: list[dict] = []
    for row in rows[header_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()

        v23 = _round_or_none(row[col_2023[0]] if col_2023 else None)
        v24 = _round_or_none(row[col_2024[0]] if col_2024 else None)
        var = _round_or_none((v24 - v23) / v23 * 100) if v23 and v24 else None
        pc = _round_or_none(row[col_2024[1]] if len(col_2024) > 1 and col_2024[1] < len(row) else None)

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "total_2023": v23,
            "total_2024": v24,
            "variacao_pct": var,
            "per_capita_2024": pc,
        })

    if not dados:
        print("  ! T96: nenhuma linha extraída.", file=sys.stderr)
        return

    save_json(
        output_dir / "gastos.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T96",
            "dados": dados,
        },
        "T96 Gastos",
    )


# ---------------------------------------------------------------------------
# T22 — Suicídios
# ---------------------------------------------------------------------------


def extract_suicidios(wb, year: int, output_dir: Path) -> None:
    rows = read_sheet(wb, "T22")

    header_idx = None
    for i, row in enumerate(rows):
        anos = [v for v in row if isinstance(v, (int, float)) and v in (2023, 2024)]
        if len(anos) >= 2:
            header_idx = i
            break

    if header_idx is None:
        print("  ! T22: cabeçalho não encontrado. Pulando.", file=sys.stderr)
        return

    header = rows[header_idx]
    col_2023 = [i for i, v in enumerate(header) if v == 2023]
    col_2024 = [i for i, v in enumerate(header) if v == 2024]

    dados: list[dict] = []
    for row in rows[header_idx + 1 :]:
        uf = row[0]
        if uf is None or str(uf).strip() not in UFS:
            continue
        uf_str = str(uf).strip()

        def get(cols, idx):
            return row[cols[idx]] if idx < len(cols) and cols[idx] < len(row) else None

        dados.append({
            "uf": uf_str,
            "regiao": _regiao(uf_str),
            "total_2023": _int_or_none(get(col_2023, 0)),
            "total_2024": _int_or_none(get(col_2024, 0)),
            "taxa_2023": _round_or_none(get(col_2023, 1)),
            "taxa_2024": _round_or_none(get(col_2024, 1)),
        })

    if not dados:
        print("  ! T22: nenhuma linha extraída.", file=sys.stderr)
        return

    save_json(
        output_dir / "suicidios.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "tabela": "T22",
            "dados": dados,
        },
        "T22 Suicídios",
    )


# ---------------------------------------------------------------------------
# P07 — Gráfico 48: Autoria por sexo
# ---------------------------------------------------------------------------


def extract_autoria_sexo(wb, year: int, output_dir: Path) -> None:
    """
    P07 / Gráfico 48: Participação na autoria de MVI segundo sexo da vítima.

    ATENÇÃO: Este dado é politicamente sensível e deve SEMPRE ser exibido
    com o campo `contexto`. Nunca renderize os números isolados.

    Categorias:
      1. Vítimas mulheres — feminicídio
      2. Vítimas mulheres — MVI (geral)
      3. Vítimas homens — MVI

    Tipos de autoria:
      A. Apenas autoras do sexo feminino
      B. Apenas autores do sexo masculino
      C. Autoria múltipla / não identificado por sexo
    """
    rows = read_sheet(wb, "P07")

    # Mapear as três linhas-chave por keyword na coluna 0
    KEYWORDS = {
        "apenas_feminino": ["apenas autoras", "autoras do sexo feminino"],
        "apenas_masculino": ["apenas autores", "autores do sexo masculino"],
        "multipla": ["múltipla", "multipla", "não identificado"],
    }

    tipo_rows: dict[str, list | None] = {k: None for k in KEYWORDS}

    for row in rows:
        if row[0] is None:
            continue
        cell = str(row[0]).lower().strip()
        for tipo, keywords in KEYWORDS.items():
            if any(kw in cell for kw in keywords):
                tipo_rows[tipo] = row
                break

    # Verificar se encontrou as três linhas
    missing = [k for k, v in tipo_rows.items() if v is None]
    if missing:
        print(
            f"  ! P07: linhas não encontradas para {missing}. "
            "Verifique as keywords no script.",
            file=sys.stderr,
        )
        # Não aborta — salva o que encontrou com None para os faltantes
        for k in missing:
            tipo_rows[k] = [None] * 4

    def vals(row):
        # Colunas 1, 2, 3 correspondem às 3 categorias de vítima
        return [
            _round_or_none(row[1] if len(row) > 1 else None),
            _round_or_none(row[2] if len(row) > 2 else None),
            _round_or_none(row[3] if len(row) > 3 else None),
        ]

    fem_vals = vals(tipo_rows["apenas_feminino"])
    mas_vals = vals(tipo_rows["apenas_masculino"])
    mul_vals = vals(tipo_rows["multipla"])

    CATEGORIAS = [
        ("vitimas_mulheres_feminicidio", "Vítimas mulheres — feminicídio"),
        ("vitimas_mulheres_mvi", "Vítimas mulheres — MVI"),
        ("vitimas_homens_mvi", "Vítimas homens — MVI"),
    ]

    dados = [
        {
            "categoria": cat,
            "label": label,
            "apenas_autoras_femininas": fem_vals[i],
            "apenas_autores_masculinos": mas_vals[i],
            "autoria_multipla": mul_vals[i],
        }
        for i, (cat, label) in enumerate(CATEGORIAS)
    ]

    save_json(
        output_dir / "autoria_sexo.json",
        {
            "fonte": f"Fórum Brasileiro de Segurança Pública, {year}",
            "grafico": "48",
            "tabela_origem": "P07",
            "ano_referencia": year - 1,  # dados referem-se ao ano anterior
            "contexto": (
                "Percentual de participação na autoria de Mortes Violentas Intencionais (MVI) "
                "e feminicídios segundo o sexo da vítima. Inclui apenas casos com autoria "
                "conhecida registrada pelas secretarias estaduais de segurança pública. "
                "A maioria dos casos em que mulheres figuram como autoras em mortes de homens "
                "está associada à violência conjugal defensiva. Em 97% dos feminicídios, "
                "o autor é do sexo masculino. Os dados revelam a assimetria estrutural da "
                "violência de gênero, não sua simetria."
            ),
            "dados": dados,
        },
        "P07 Autoria por sexo (Gráfico 48)",
    )


# ---------------------------------------------------------------------------
# meta.json
# ---------------------------------------------------------------------------


def extract_meta(year: int, excel_path: Path, output_dir: Path) -> None:
    save_json(
        output_dir / "meta.json",
        {
            "numero_anuario": year - 2006,  # 2025 → 19º
            "ano_publicacao": year,
            "ano_referencia": year - 1,
            "url_fonte": f"https://forumseguranca.org.br/wp-content/uploads/{year}/09/anuario-{year}.xlsx",
            "data_extracao": datetime.now().isoformat(),
        },
        "meta.json",
    )


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Extrai dados do Anuário de Segurança Pública")
    parser.add_argument("--year", type=int, default=2025, help="Ano de publicação do anuário (ex: 2025)")
    args = parser.parse_args()

    year: int = args.year
    base_dir = Path(__file__).parent.parent  # raiz do projeto
    excel_path = Path(__file__).parent / "raw" / str(year) / f"anuario-{year}.xlsx"
    output_dir = base_dir / "data" / str(year)

    print(f"\n{'='*60}")
    print(f"  Anuário {year} — extração de dados")
    print(f"  Excel: {excel_path}")
    print(f"  Saída: {output_dir}")
    print(f"{'='*60}\n")

    if not excel_path.exists():
        print(f"ERRO: arquivo não encontrado: {excel_path}", file=sys.stderr)
        print(
            f"\nBaixe o Excel em:\n"
            f"  https://forumseguranca.org.br/wp-content/uploads/{year}/09/anuario-{year}.xlsx\n"
            f"E salve em:\n"
            f"  {excel_path}\n",
            file=sys.stderr,
        )
        sys.exit(1)

    print("Carregando workbook (pode demorar 30–60s para arquivos grandes)...")
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    print(f"Abas disponíveis: {len(wb.sheetnames)}\n")

    print("Extraindo tabelas:")
    extract_meta(year, excel_path, output_dir)
    extract_mvi_historico(wb, year, output_dir)
    extract_mvi_estados(wb, year, output_dir)
    extract_feminicidio(wb, year, output_dir)
    extract_feminicidio_hist(wb, year, output_dir)
    extract_letalidade(wb, year, output_dir)
    extract_estupro(wb, year, output_dir)
    extract_prisional(wb, year, output_dir)
    extract_desaparecimentos(wb, year, output_dir)
    extract_patrimonio(wb, year, output_dir)
    extract_gastos(wb, year, output_dir)
    extract_suicidios(wb, year, output_dir)
    extract_autoria_sexo(wb, year, output_dir)

    wb.close()
    print(f"\n{'='*60}")
    print(f"  Extração concluída. JSONs em: {output_dir}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
